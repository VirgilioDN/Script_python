#Sistema
import pyautogui
import os
import time
import random
import urllib
import pydub
from pydub import AudioSegment
import shutil

#selenium 
from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

#transcrção de audio
import speech_recognition as sr

#scraping
from bs4 import BeautifulSoup

#planilha
from openpyxl import load_workbook

#caminhos necessarios para funcionamento (importante alterar!!)
firefox_path = r"C:\Program Files\Mozilla Firefox\firefox.exe" #altere para o caminho no seu pc
geckodriver_path = r"C:\Users\virgilio\Downloads\geckodriver-v0.36.0-win64\geckodriver.exe" #altere para o caminho no seu pc
ffmpeg_bin = r"C:\Users\virgilio\Downloads\ffmpeg-master-latest-win64-gpl-shared\ffmpeg-master-latest-win64-gpl-shared\bin" #altere para o caminho no seu pc
if ffmpeg_bin not in os.environ.get("PATH", ""):
    os.environ["PATH"] += os.pathsep + ffmpeg_bin

AudioSegment.converter = os.path.join(ffmpeg_bin, "ffmpeg.exe")
AudioSegment.ffprobe = os.path.join(ffmpeg_bin, "ffprobe.exe")

path_to_download = os.path.join(os.getcwd(), "audio")
os.makedirs(path_to_download, exist_ok=True)

#configuração dos drivers do firefox
options = Options()
options.binary_location = firefox_path
options.set_preference('browser.download.folderList', 2)
options.set_preference('browser.helperApps.alwaysAsk.force', False)
options.set_preference('browser.download.manager.showWhenStarting', False)
options.set_preference('browser.helperApps.neverAsk.saveToDisk', 'application/zip')

service = Service(geckodriver_path)
driver = webdriver.Firefox(service=service, options=options)

#redimensiona a janela
width = height = 800
ss_w, ss_h = pyautogui.size()
driver.set_window_size(width, height)
driver.set_window_position(ss_w / 2 - width / 2, ss_h / 2 - height / 2)
print('[INFO] Firefox: aberto com sucesso')


#carrega a planilha
wb = load_workbook(filename="link_professores.xlsx") #ALTERE O NOME DA PLANILHA AQUI
sheet_name = "Página1"
if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
else:
    ws = wb.create_sheet(sheet_name)

cont_ws = ws 

if cont_ws.cell(row=1, column=1).value is None:
    cont_ws.cell(row=1, column=1, value="Nome")
if cont_ws.cell(row=1, column=2).value is None:
    cont_ws.cell(row=1, column=2, value="Link")

#função auxilar para inserir ano
def get_col_for_year_insert(planilha, ano):
    try:
        ano_int = int(str(ano).strip())
    except Exception:
        nova_col = planilha.max_column + 1
        planilha.cell(row=1, column=nova_col, value=str(ano))
        return nova_col

    if planilha.max_column < 3:
        col = 3
        planilha.cell(row=1, column=col, value=str(ano_int))
        return col

    for col_index in range(3, planilha.max_column + 1):
        hdr = planilha.cell(row=1, column=col_index).value
        try:
            hdr_int = int(str(hdr).strip())
        except Exception:
            continue
        if hdr_int == ano_int:
            return col_index
        if ano_int > hdr_int:
            planilha.insert_cols(col_index)
            planilha.cell(row=1, column=col_index, value=str(ano_int))
            return col_index

    #se não encontrou posição adequada, adiciona no final
    nova_col = planilha.max_column + 1
    planilha.cell(row=1, column=nova_col, value=str(ano_int))
    return nova_col

pausa = 0.3

#loop 
for lin_index, linha in enumerate(cont_ws.iter_rows(min_row=2, values_only=False), start=2):
    nome_cell = linha[0]
    link_cell = linha[1]
    
    #verifica se o o link ja foi processado
    processed = False
    for col in range(3, cont_ws.max_column + 1):
        val = cont_ws.cell(row=lin_index, column=col).value
        if val is not None and str(val).strip() != "":
            processed = True
            break

    if processed:
        print(f"[INFO] Pulando {nome_cell.value} - já processado.")
        continue

    nome = nome_cell.value
    link = link_cell.value
    driver.get(link)
    print(f"[INFO] Carregando página de {nome}")

    if random.random() < pausa: 
        print("[INFO] Pausa aleatória ativada. Aguardando que uma pessoa resolva o captcha")

        start_wait = time.time()
        max_wait_seconds = 60  # timeout máximo
        while True:
            # 1) Verifica se a página já avançou (elemento de artigos presente)
            try:
                if len(driver.find_elements(By.ID, "artigos-completos")) > 0:
                    print("[INFO] Detecção automática: 'artigos-completos' encontrado. Prosseguindo.")
                    pausa = 0.3
                    break
            except Exception:
                pass

            # 2) Verifica se o botão submit existe e está habilitado (sinal que captcha foi resolvido)
            try:
                btns = driver.find_elements(By.ID, "submitBtn")
                if btns:
                    try:
                        if btns[0].is_enabled():
                            pausa = 0.3
                            break
                    except Exception:
                        pass
            except Exception:
                pass

            #timeout
            if time.time() - start_wait > max_wait_seconds:
                print(f"[WARN] Timeout de espera ({max_wait_seconds}s) atingido. Prosseguindo automaticamente.")
                break
            
            time.sleep(1)  # espera 1s antes de checar de novo
    else:
        pausa += 0.04
        #burlar captcha
        # Find iframe tag and switch to that iframe context
        frames = driver.find_elements(By.TAG_NAME, 'iframe')
        driver.switch_to.frame(frames[0])

        # Click on recaptcha checkbox and switch to default context
        driver.find_element(By.CLASS_NAME, 'recaptcha-checkbox-border').click()
        driver.switch_to.default_content()

        # Investigate submit button
        button = driver.find_element(By.ID, 'submitBtn')
        time.sleep(random.randint(1, 2))

        if not button.is_enabled():
            try:
                # Find iframe tag and switch to that iframe context
                frames = driver.find_element(By.XPATH, '/html/body/div[2]/div[4]').find_elements(By.TAG_NAME, 'iframe')
                driver.switch_to.frame(frames[0])

                # Click on recaptcha audio button (alternative way to solve recaptcha)
                time.sleep(random.randint(1, 2))
                driver.find_element(By.ID, 'recaptcha-audio-button').click()

                # Switch to default context again
                driver.switch_to.default_content()

                # Find iframe tag and switch to the last context
                frames = driver.find_elements(By.TAG_NAME, 'iframe')
                driver.switch_to.frame(frames[-1])

                # [Optional] Wait 1 second and play audio
                time.sleep(1)
                driver.find_element(By.XPATH, '/html/body/div/div/div[3]/div/button').click()

                #================================================#
                # From now on: download the mp3 audio source,
                # convert to wav format,
                # feed speech recognition algorithm,
                # translate to string,
                # and send string back to recaptcha frame
                #================================================#

                # Download mp3 file
                src = driver.find_element(By.ID, 'audio-source').get_attribute('src')
                file_name = path_to_download + '/sample.mp3'
                urllib.request.urlretrieve(src, file_name)

                try:
                    # Get file and convert to wav extension
                    sound = pydub.AudioSegment.from_mp3(file_name)
                    file_name = file_name.replace('.mp3', '.wav')
                    sound.export(file_name, format = 'wav')
                except Exception:
                    print("não foi possivel transcrever o audio, execute novamente para recomeçar")
                    raise Exception

                # Submit audio to a speechrecognition algorithm from Google
                sample_audio = sr.AudioFile(file_name)
                r = sr.Recognizer()
                with sample_audio as source:
                    audio = r.record(source)

                key = r.recognize_google(audio)

                # Send string (key) back to recaptcha page and switch to default context again
                driver.find_element(By.ID, 'audio-response').send_keys(key.lower())
                driver.find_element(By.ID, 'audio-response').send_keys(Keys.ENTER)
                driver.switch_to.default_content()

                # Submit solution by clicking the button
                time.sleep(1)
                driver.find_element(By.ID, 'submitBtn').click()
            except Exception:
                print("algum erro ocorreu. tente novamente ou aguarde alguns instantes")
                raise Exception
        else:
            button.click()

    wait = WebDriverWait(driver, 10)
    pag_artigos = wait.until(EC.presence_of_element_located((By.ID, "artigos-completos")))
    artigos = driver.find_element(By.ID, "artigos-completos")
    html_artigo = artigos.get_attribute('outerHTML')

    soup = BeautifulSoup(html_artigo, 'html.parser')
    anos = [span.get_text(strip=True) for span in soup.find_all("span", {"data-tipo-ordenacao": "ano"})]

    contagem = {}
    for ano in anos:
        contagem[ano] = contagem.get(ano, 0) + 1

    cont_ws.cell(row=lin_index, column=1, value=nome)
    cont_ws.cell(row=lin_index, column=2, value=link)
    for ano, qtd in contagem.items():
        col = get_col_for_year_insert(cont_ws, ano)
        cont_ws.cell(row=lin_index, column=col, value=qtd)

    wb.save("link_professores.xlsx") #ALTERE O NOME DA PLANILHA AQUI


driver.quit()
